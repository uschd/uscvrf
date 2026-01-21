import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# ================= CONFIGURATION =================
FILE_PATH = 'Test2.xlsx'       # Main Project File
REF_FILE_PATH = 'mitsubishi_electric_vrf.xlsx' # The file with Model/TR data
INPUT_SHEET = 'Input'
SUMMARY_SHEET = 'Summary'

# ================= HELPER FUNCTIONS =================

def get_floor_mapping():
    return {
        -1: "BASEMENT", 0: "GROUND FLOOR", 1: "FIRST FLOOR", 2: "SECOND FLOOR",
        3: "THIRD FLOOR", 4: "FOURTH FLOOR", 5: "FIFTH FLOOR", 6: "SIXTH FLOOR",
        7: "SEVENTH FLOOR", 8: "EIGHTH FLOOR", 9: "NINTH FLOOR", 10: "TENTH FLOOR"
    }

def clean_tonnage_value(val):
    try:
        if pd.isna(val): return None
        s_val = str(val).lower().replace('tr', '').strip()
        s_val = s_val.split()[0]
        return float(s_val)
    except:
        return None

def load_reference_data(ref_path):
    """
    Robust loader for the Reference file (handles empty rows/mixed data).
    """
    if not os.path.exists(ref_path):
        print(f"Warning: Reference file '{ref_path}' not found.")
        return pd.DataFrame()

    try:
        df = pd.read_excel(ref_path, header=None)
        clean_data = []
        
        for index, row in df.iterrows():
            if len(row) < 4: continue # Skip short rows

            col0_val = row.iloc[0]
            if pd.isna(col0_val): continue
                
            col0_str = str(col0_val).strip()
            
            # Check for '1' or '1.0' in the first column
            if col0_str == '1' or col0_str == '1.0':
                tr_raw = row.iloc[2]
                model_val = row.iloc[3]
                
                tr_num = clean_tonnage_value(tr_raw)
                
                if tr_num is not None and pd.notna(model_val):
                    clean_data.append({
                        'TR_Numeric': tr_num,
                        'Model': str(model_val).strip()
                    })
        
        return pd.DataFrame(clean_data)
    except Exception as e:
        print(f"Error loading reference data: {e}")
        return pd.DataFrame()

def find_closest_model(target_tr, ref_df):
    if ref_df.empty: return None, None
    
    ref_df = ref_df.copy()
    ref_df['diff'] = (ref_df['TR_Numeric'] - target_tr).abs()
    
    closest_idx = ref_df['diff'].idxmin()
    return ref_df.loc[closest_idx, 'TR_Numeric'], ref_df.loc[closest_idx, 'Model']

# ================= MAIN PROCESSING =================

def fill_detailed_data(input_df, ref_df):
    mapper = get_floor_mapping()
    data_rows = []
    
    # Iterate through input rows
    for index, row in input_df.iterrows():
        # 1. READ COLUMNS (Safe Method)
        
        # Column 0: Floor
        try:
            val_0 = row.iloc[0]
            floor_idx = int(val_0) if pd.notna(val_0) else 0
        except:
            continue # Skip header rows or bad data

        # Column 1: Description
        try:
            description = row.iloc[1] if pd.notna(row.iloc[1]) else ""
        except:
            description = ""

        # Column 2: Area (The fix is here)
        try:
            # Check if Column 2 exists in this row
            if len(row) > 2:
                val_2 = row.iloc[2]
                area = float(val_2) if pd.notna(val_2) else 0
            else:
                area = 0 # Default to 0 if column is missing
        except:
            area = 0
            
        floor_name = mapper.get(floor_idx, f"FLOOR {floor_idx}")
        
        # 2. Logic Calculations
        tonnage_req = area / 110.0
        selected_tr, model_name = find_closest_model(tonnage_req, ref_df)
        
        qty = 1
        unit_type = "Mid Static Ducted Unit"
        total_tonnage = selected_tr * qty if selected_tr else 0

        data_rows.append({
            'FLOOR': floor_name,
            'DESCRIPTIONS': description,
            'AREA IN SQ.FT': int(round(area)),
            'TONNAGE REQUIRED': round(tonnage_req, 1),
            'SELECTED TR FOR AREA': selected_tr,
            'QTY (Nos.)': qty,
            'Total Tonnage': total_tonnage,
            'Model': model_name,
            'UNIT TYPE': unit_type,
            'Cap.Index': None
        })
        
    return pd.DataFrame(data_rows)

def apply_styling(file_path, sheet_name):
    wb = load_workbook(file_path)
    if sheet_name not in wb.sheetnames: return
    ws = wb[sheet_name]
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    peach_fill = PatternFill(start_color="F4C7AA", end_color="F4C7AA", fill_type="solid")
    
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "BASIS OF DESIGN"
    title_cell.font = Font(bold=True, color="FF0000")
    title_cell.alignment = Alignment(horizontal='center')
    
    for col in range(1, 11):
        cell = ws.cell(row=2, column=col)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=10):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if cell.column == 2:
                cell.alignment = Alignment(horizontal='left', vertical='center')

    for row in range(3, ws.max_row + 1):
        cell = ws.cell(row=row, column=1)
        cell.fill = peach_fill
        cell.font = Font(bold=True, size=9)

    widths = [15, 30, 10, 10, 10, 8, 10, 18, 25, 10]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

    wb.save(file_path)

def main():
    if not os.path.exists(FILE_PATH):
        print(f"Error: Main file '{FILE_PATH}' not found.")
        return

    # 1. Read Input Data (RESTORED TO SIMPLE VERSION)
    try:
        print(f"Reading '{FILE_PATH}'...")
        # We REMOVE 'usecols=[0,1,2]' because that causes the crash if Col 2 is empty.
        # We read everything, then pick what we need inside the loop.
        df_input = pd.read_excel(FILE_PATH, sheet_name=INPUT_SHEET, header=None)
        
    except Exception as e:
        print(f"Error reading input sheet: {e}")
        return

    # 2. Read Reference Data
    print("Reading reference database...")
    df_ref = load_reference_data(REF_FILE_PATH)
    
    if df_ref.empty:
        print("Error: Could not extract valid data from Reference file.")

    # 3. Process Data
    print("Processing logic...")
    df_summary = fill_detailed_data(df_input, df_ref)

    # 4. Write to Excel
    try:
        with pd.ExcelWriter(FILE_PATH, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            if SUMMARY_SHEET in writer.book.sheetnames:
                print(f"Updating '{SUMMARY_SHEET}'...")
                idx = writer.book.sheetnames.index(SUMMARY_SHEET)
                writer.book.remove(writer.book.worksheets[idx])
            else:
                print(f"Creating '{SUMMARY_SHEET}'...")
            
            df_summary.to_excel(writer, sheet_name=SUMMARY_SHEET, index=False)
                
    except Exception as e:
        print(f"Error writing to Excel: {e}")
        return

    # 5. Apply Styling
    apply_styling(FILE_PATH, SUMMARY_SHEET)
    print("Process Complete.")

if __name__ == "__main__":
    main{}
